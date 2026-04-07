#!/usr/bin/env python3
"""
Phase 1C: Retrieve abstracts from PubMed for all publications.
Uses NCBI E-utilities (esearch to find PMID from DOI, efetch to get abstract).
Updates ABCD_SM_Review_Codebook.xlsx with tab 1C_Abstract_Retrieval.
Then re-runs keyword screening on abstracts and updates tab 1B.
"""

import json
import csv
import time
import re
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

import requests
from openpyxl import load_workbook
from openpyxl.styles import Font

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
RAW_DIR = DATA_DIR / "raw"

HEADERS = {
    "User-Agent": "ABCD-Review-Bot/1.0 (mailto:daniel.zweben@temple.edu)"
}

KEYWORDS = [
    "screen time", "social media", "smartphone", "mobile phone",
    "video game", "gaming", "texting", "digital media",
    "internet use", "problematic screen", "screen use",
    "television", "media use",
]

def keyword_col_name(kw):
    return "kw_" + kw.replace(" ", "_")

def doi_to_pmid(doi):
    """Look up PMID from DOI via PubMed esearch."""
    if not doi:
        return None, "no_doi"
    url = f"https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?db=pubmed&term={quote(doi)}[DOI]&retmode=json"
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
        data = resp.json()
        ids = data.get("esearchresult", {}).get("idlist", [])
        if ids:
            return ids[0], "found"
        return None, "not_found"
    except Exception as e:
        return None, f"error: {str(e)[:80]}"

def pmid_from_url(url):
    """Extract PMID from PubMed URL."""
    match = re.search(r'pubmed[./](\d+)', url)
    if match:
        return match.group(1)
    return None

def fetch_abstracts_batch(pmids):
    """Fetch abstracts for a batch of PMIDs via efetch."""
    if not pmids:
        return {}

    pmid_str = ",".join(pmids)
    url = f"https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi?db=pubmed&id={pmid_str}&retmode=xml"

    try:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        root = ET.fromstring(resp.content)

        results = {}
        for article in root.findall(".//PubmedArticle"):
            pmid_elem = article.find(".//PMID")
            if pmid_elem is None:
                continue
            pmid = pmid_elem.text

            # Get abstract
            abstract_parts = []
            for abs_text in article.findall(".//AbstractText"):
                label = abs_text.get("Label", "")
                text = "".join(abs_text.itertext()).strip()
                if label:
                    abstract_parts.append(f"{label}: {text}")
                else:
                    abstract_parts.append(text)

            abstract = " ".join(abstract_parts)
            results[pmid] = abstract

        return results
    except Exception as e:
        print(f"  Error fetching batch: {e}")
        return {}

def main():
    # Load publications
    with open(RAW_DIR / "all_publications.json") as f:
        publications = json.load(f)

    print(f"Processing {len(publications)} publications for abstract retrieval")

    # Step 1: Resolve all DOIs to PMIDs
    print("\nStep 1: Resolving DOIs to PMIDs...")
    pmid_map = {}  # row_id -> {pmid, status}

    # Check for checkpoint
    checkpoint_path = RAW_DIR / "pmid_checkpoint.json"
    if checkpoint_path.exists():
        with open(checkpoint_path) as f:
            pmid_map = json.load(f)
        print(f"  Loaded checkpoint with {len(pmid_map)} entries")

    for i, pub in enumerate(publications):
        row_id = str(pub["row_id"])
        if row_id in pmid_map:
            continue

        # First check if URL is already a PubMed link
        pmid = pmid_from_url(pub.get("url", ""))
        if pmid:
            pmid_map[row_id] = {"pmid": pmid, "status": "found"}
        else:
            doi = pub.get("doi", "")
            pmid, status = doi_to_pmid(doi)
            pmid_map[row_id] = {"pmid": pmid, "status": status}
            time.sleep(0.35)  # Rate limit: ~3/sec

        if (i + 1) % 50 == 0:
            found = sum(1 for v in pmid_map.values() if v["pmid"])
            print(f"  Processed {i+1}/{len(publications)} ({found} PMIDs found)")
            # Save checkpoint
            with open(checkpoint_path, "w") as f:
                json.dump(pmid_map, f)

    # Final checkpoint
    with open(checkpoint_path, "w") as f:
        json.dump(pmid_map, f)

    found_count = sum(1 for v in pmid_map.values() if v["pmid"])
    print(f"\n  PMID resolution complete: {found_count}/{len(publications)} found")

    # Step 2: Batch fetch abstracts
    print("\nStep 2: Fetching abstracts in batches...")
    all_pmids = [v["pmid"] for v in pmid_map.values() if v["pmid"]]
    abstract_map = {}  # pmid -> abstract text

    # Check for abstract checkpoint
    abs_checkpoint_path = RAW_DIR / "abstract_checkpoint.json"
    if abs_checkpoint_path.exists():
        with open(abs_checkpoint_path) as f:
            abstract_map = json.load(f)
        print(f"  Loaded abstract checkpoint with {len(abstract_map)} entries")

    # Only fetch PMIDs we don't have yet
    remaining_pmids = [p for p in all_pmids if p not in abstract_map]
    batch_size = 200

    for i in range(0, len(remaining_pmids), batch_size):
        batch = remaining_pmids[i:i+batch_size]
        results = fetch_abstracts_batch(batch)
        abstract_map.update(results)

        # Mark PMIDs with no abstract
        for pmid in batch:
            if pmid not in abstract_map:
                abstract_map[pmid] = ""

        print(f"  Batch {i//batch_size + 1}: fetched {len(results)} abstracts ({len(batch)} PMIDs)")
        time.sleep(0.5)

        # Checkpoint
        with open(abs_checkpoint_path, "w") as f:
            json.dump(abstract_map, f)

    has_abstract = sum(1 for v in abstract_map.values() if v)
    print(f"\n  Abstract retrieval complete: {has_abstract}/{len(all_pmids)} have abstracts")

    # Step 3: Build records for tab 1C
    print("\nStep 3: Building abstract retrieval records...")
    records = []
    for pub in publications:
        row_id = str(pub["row_id"])
        pmid_info = pmid_map.get(row_id, {"pmid": None, "status": "missing"})
        pmid = pmid_info["pmid"]

        abstract = abstract_map.get(pmid, "") if pmid else ""

        if pmid and abstract:
            abs_status = "retrieved"
        elif pmid and not abstract:
            abs_status = "unavailable"
        else:
            abs_status = "no_pmid"

        records.append({
            "row_id": pub["row_id"],
            "doi": pub["doi"],
            "pmid": pmid or "",
            "pmid_lookup_status": pmid_info["status"],
            "abstract_text": abstract,
            "abstract_retrieval_status": abs_status,
            "abstract_word_count": len(abstract.split()) if abstract else 0,
            "coder": "AI",
            "date_coded": datetime.now().strftime("%Y-%m-%d"),
            "notes": "",
        })

    # Step 4: Re-screen abstracts against keywords
    print("\nStep 4: Re-screening abstracts against keywords...")
    # Load existing keyword screening
    with open(RAW_DIR / "keyword_screening.csv") as f:
        reader = csv.DictReader(f)
        kw_records = list(reader)

    # Create lookup
    kw_by_row = {int(r["row_id"]): r for r in kw_records}

    abstract_only_matches = 0
    for rec in records:
        row_id = rec["row_id"]
        abstract = rec["abstract_text"].lower()
        kw_rec = kw_by_row.get(row_id)
        if not kw_rec:
            continue

        if not abstract:
            continue

        abs_hit_count = 0
        title_had_hit = kw_rec["any_keyword_hit"] in (True, "True", "1", 1)

        for kw in KEYWORDS:
            col = keyword_col_name(kw)
            if kw.lower() in abstract:
                abs_hit_count += 1
                # If title didn't have this keyword but abstract does, update
                if kw_rec[col] in (0, "0", False, "False"):
                    kw_rec[col] = 1  # Mark as hit (found in abstract)

        kw_rec["abstract_hit_count"] = abs_hit_count

        if abs_hit_count > 0 and not title_had_hit:
            kw_rec["any_keyword_hit"] = True
            kw_rec["match_location"] = "abstract"
            kw_rec["include_exclude"] = "INCLUDE"
            abstract_only_matches += 1
        elif abs_hit_count > 0 and title_had_hit:
            kw_rec["match_location"] = "both"

        kw_rec["title_hit_count"] = sum(
            1 for kw in KEYWORDS
            if kw.lower() in kw_rec.get("title", "").lower()
        )

    total_included = sum(1 for r in kw_records if r.get("include_exclude") == "INCLUDE" or r.get("any_keyword_hit") in (True, "True", 1, "1"))
    print(f"  Additional abstract-only matches: {abstract_only_matches}")
    print(f"  Total included after abstract screening: {total_included}")

    # Step 5: Save everything to Excel
    print("\nStep 5: Saving to Excel...")
    wb_path = DATA_DIR / "ABCD_SM_Review_Codebook.xlsx"
    wb = load_workbook(wb_path)

    # Tab 1C
    if "1C_Abstract_Retrieval" in wb.sheetnames:
        del wb["1C_Abstract_Retrieval"]
    ws_1c = wb.create_sheet("1C_Abstract_Retrieval")

    headers_1c = list(records[0].keys())
    ws_1c.append(headers_1c)
    for cell in ws_1c[1]:
        cell.font = Font(bold=True)
    for rec in records:
        ws_1c.append([rec[h] for h in headers_1c])

    # Update tab 1B with abstract results
    if "1B_Keyword_Screening" in wb.sheetnames:
        del wb["1B_Keyword_Screening"]
    ws_1b = wb.create_sheet("1B_Keyword_Screening")

    headers_1b = list(kw_records[0].keys())
    ws_1b.append(headers_1b)
    for cell in ws_1b[1]:
        cell.font = Font(bold=True)
    for rec in kw_records:
        ws_1b.append([rec.get(h, "") for h in headers_1b])

    wb.save(wb_path)
    print(f"  Updated workbook with tabs 1C_Abstract_Retrieval and 1B_Keyword_Screening")

    # Save updated keyword screening CSV
    with open(RAW_DIR / "keyword_screening_with_abstracts.csv", "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers_1b)
        writer.writeheader()
        writer.writerows(kw_records)

    # Save abstract records CSV
    with open(RAW_DIR / "abstract_retrieval.csv", "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers_1c)
        writer.writeheader()
        writer.writerows(records)

    print("\nDone!")
    return records, kw_records

if __name__ == "__main__":
    main()
